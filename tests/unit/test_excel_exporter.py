"""Tests pour exporters/excel.py — export Excel et résumé console."""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl

from compta_ecom.config.loader import AppConfig
from compta_ecom.exporters.excel import ANOMALIES_COLUMNS, ENTRIES_COLUMNS, export, print_summary
from compta_ecom.models import AccountingEntry, Anomaly


def _make_entry(**overrides: object) -> AccountingEntry:
    """Helper pour construire un AccountingEntry."""
    defaults: dict[str, object] = {
        "date": datetime.date(2026, 1, 15),
        "journal": "VE",
        "account": "70701250",
        "label": "Vente #001 Shopify",
        "debit": 0.0,
        "credit": 120.0,
        "piece_number": "#001",
        "lettrage": "#001",
        "channel": "shopify",
        "entry_type": "sale",
    }
    defaults.update(overrides)
    return AccountingEntry(**defaults)  # type: ignore[arg-type]


def _make_anomaly(**overrides: object) -> Anomaly:
    """Helper pour construire une Anomaly."""
    defaults: dict[str, object] = {
        "type": "orphan_sale",
        "severity": "warning",
        "reference": "#999",
        "channel": "shopify",
        "detail": "Vente sans transaction",
        "expected_value": None,
        "actual_value": None,
    }
    defaults.update(overrides)
    return Anomaly(**defaults)  # type: ignore[arg-type]


class TestExportNominal:
    """Tests de l'export Excel nominal."""

    def test_two_sheets_created(self, tmp_path: Path, sample_config: AppConfig) -> None:
        """Export crée 2 onglets : Écritures et Anomalies."""
        entries = [_make_entry(), _make_entry(piece_number="#002")]
        anomalies = [_make_anomaly()]
        output = tmp_path / "output.xlsx"

        export(entries, anomalies, output, sample_config)

        wb = openpyxl.load_workbook(output)
        assert "Écritures" in wb.sheetnames
        assert "Anomalies" in wb.sheetnames
        wb.close()

    def test_entries_columns_order(self, tmp_path: Path, sample_config: AppConfig) -> None:
        """Les colonnes de l'onglet Écritures suivent l'ordre fixe."""
        entries = [_make_entry()]
        output = tmp_path / "output.xlsx"

        export(entries, [], output, sample_config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Écritures"]
        headers = [cell.value for cell in ws[1]]
        assert headers == ENTRIES_COLUMNS
        wb.close()

    def test_entries_row_count(self, tmp_path: Path, sample_config: AppConfig) -> None:
        """Nombre de lignes = nombre d'écritures."""
        entries = [_make_entry(), _make_entry(piece_number="#002"), _make_entry(piece_number="#003")]
        output = tmp_path / "output.xlsx"

        export(entries, [], output, sample_config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Écritures"]
        data_rows = list(ws.iter_rows(min_row=2))
        assert len(data_rows) == 3
        wb.close()

    def test_dates_are_datetime(self, tmp_path: Path, sample_config: AppConfig) -> None:
        """Les dates sont écrites en datetime, pas en string."""
        entries = [_make_entry()]
        output = tmp_path / "output.xlsx"

        export(entries, [], output, sample_config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Écritures"]
        date_cell = ws.cell(row=2, column=1)
        assert isinstance(date_cell.value, datetime.datetime)
        wb.close()


class TestExportAnomaliesEmpty:
    """Tests onglet Anomalies vide."""

    def test_anomalies_headers_present(self, tmp_path: Path, sample_config: AppConfig) -> None:
        """Onglet Anomalies avec en-têtes même si aucune anomalie."""
        output = tmp_path / "output.xlsx"

        export([], [], output, sample_config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Anomalies"]
        headers = [cell.value for cell in ws[1]]
        assert headers == ANOMALIES_COLUMNS
        data_rows = list(ws.iter_rows(min_row=2))
        assert len(data_rows) == 0
        wb.close()


class TestExportAnomaliesPopulated:
    """Tests onglet Anomalies peuplé."""

    def test_anomalies_data(self, tmp_path: Path, sample_config: AppConfig) -> None:
        """Anomalies présentes dans l'onglet avec les bonnes colonnes."""
        anomalies = [_make_anomaly(), _make_anomaly(reference="#888", type="unknown_psp")]
        output = tmp_path / "output.xlsx"

        export([], anomalies, output, sample_config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Anomalies"]
        headers = [cell.value for cell in ws[1]]
        assert headers == ANOMALIES_COLUMNS
        data_rows = list(ws.iter_rows(min_row=2))
        assert len(data_rows) == 2
        wb.close()


class TestPrintSummary:
    """Tests du résumé console."""

    def test_summary_format(self, capsys: object) -> None:
        """print_summary affiche le format attendu."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        entries = [
            _make_entry(entry_type="sale", channel="shopify", piece_number="#001"),
            _make_entry(entry_type="sale", channel="shopify", piece_number="#001", account="4457250"),
            _make_entry(entry_type="settlement", channel="shopify", piece_number="#001"),
            _make_entry(entry_type="payout", channel="shopify", piece_number="P001"),
        ]
        anomalies = [_make_anomaly()]
        channel_errors: list[tuple[str, str]] = [("manomano", "Colonnes manquantes")]

        print_summary(entries, anomalies, channel_errors)

        captured = capsys.readouterr()
        assert "Résumé" in captured.out
        assert "shopify" in captured.out
        assert "sale" in captured.out
        assert "settlement" in captured.out
        assert "payout" in captured.out
        assert "1 warning/error, 0 info" in captured.out
        assert "manomano" in captured.out
        assert "Colonnes manquantes" in captured.out

    def test_summary_no_errors(self, capsys: object) -> None:
        """Pas de section canaux en erreur si la liste est vide."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        entries = [_make_entry()]
        print_summary(entries, [], [])

        captured = capsys.readouterr()
        assert "Canaux en erreur" not in captured.out

    def test_summary_ventilation_by_type(self, capsys: object) -> None:
        """Ventilation par type : 3 anomalies de 2 types → affichage correct."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        anomalies = [
            _make_anomaly(type="tva_mismatch", severity="warning", channel="shopify"),
            _make_anomaly(type="tva_mismatch", severity="warning", channel="shopify"),
            _make_anomaly(type="amount_mismatch", severity="warning", channel="manomano"),
        ]

        print_summary([], anomalies, [])

        captured = capsys.readouterr()
        assert "tva_mismatch" in captured.out
        assert "amount_mismatch" in captured.out
        assert "3 warning/error, 0 info" in captured.out

    def test_summary_ventilation_by_channel(self, capsys: object) -> None:
        """Ventilation par canal : anomalies multi-canal → affichage correct."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        anomalies = [
            _make_anomaly(type="tva_mismatch", channel="shopify"),
            _make_anomaly(type="tva_mismatch", channel="shopify"),
            _make_anomaly(type="amount_mismatch", channel="manomano"),
        ]

        print_summary([], anomalies, [])

        captured = capsys.readouterr()
        assert "Par canal :" in captured.out
        assert "shopify" in captured.out
        assert "manomano" in captured.out

    def test_summary_info_vs_warning(self, capsys: object) -> None:
        """Distinction info/warning : mélange severity → compteurs séparés."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        anomalies = [
            _make_anomaly(type="tva_mismatch", severity="warning", channel="shopify"),
            _make_anomaly(type="missing_payout", severity="info", channel="shopify"),
            _make_anomaly(type="missing_payout", severity="info", channel="manomano"),
        ]

        print_summary([], anomalies, [])

        captured = capsys.readouterr()
        assert "1 warning/error, 2 info" in captured.out
        assert "(info)" in captured.out

    def test_summary_no_anomalies(self, capsys: object) -> None:
        """Aucune anomalie → message dédié."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        print_summary([], [], [])

        captured = capsys.readouterr()
        assert "Aucune anomalie détectée" in captured.out

    def test_summary_only_info_anomalies(self, capsys: object) -> None:
        """Anomalies uniquement info → 0 warning/error, N info."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        anomalies = [
            _make_anomaly(type="missing_payout", severity="info", channel="shopify"),
            _make_anomaly(type="missing_payout", severity="info", channel="manomano"),
        ]

        print_summary([], anomalies, [])

        captured = capsys.readouterr()
        assert "0 warning/error, 2 info" in captured.out
