"""Tests d'intégration — POST /api/download/excel."""

from __future__ import annotations

from io import BytesIO

import openpyxl


def test_download_excel_returns_xlsx(client, shopify_files):
    """Upload Shopify → fichier .xlsx valide."""
    response = client.post("/api/download/excel", files=shopify_files)
    assert response.status_code == 200

    # Vérifier que le contenu est un Excel valide
    wb = openpyxl.load_workbook(BytesIO(response.content))
    assert "Écritures" in wb.sheetnames
    assert "Anomalies" in wb.sheetnames
    wb.close()


def test_download_excel_content_type(client, shopify_files):
    """Content-Type correct pour Excel."""
    response = client.post("/api/download/excel", files=shopify_files)
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_download_excel_content_disposition(client, shopify_files):
    """Content-Disposition contient le nom de fichier avec date."""
    response = client.post("/api/download/excel", files=shopify_files)
    assert response.status_code == 200
    disposition = response.headers["content-disposition"]
    assert disposition.startswith("attachment; filename=")
    assert "ecritures-" in disposition
    assert ".xlsx" in disposition


def test_download_excel_has_entries(client, shopify_files):
    """Le fichier Excel contient des écritures."""
    response = client.post("/api/download/excel", files=shopify_files)
    wb = openpyxl.load_workbook(BytesIO(response.content))
    ws = wb["Écritures"]
    # Header + au moins une ligne de données
    assert ws.max_row > 1
    wb.close()


def test_download_excel_invalid_file(client):
    """Fichier non-CSV → 422."""
    files = [("files", ("data.txt", b"content", "text/plain"))]
    response = client.post("/api/download/excel", files=files)
    assert response.status_code == 422


def test_download_excel_no_channel(client):
    """Fichier CSV inconnu → 422."""
    files = [("files", ("unknown.csv", b"col1,col2\nval1,val2", "text/csv"))]
    response = client.post("/api/download/excel", files=files)
    assert response.status_code == 422
