"""Tests d'intégration — export_to_bytes, export_csv_to_bytes, export_anomalies_csv_to_bytes."""

from __future__ import annotations

import datetime
from io import BytesIO

import openpyxl
import pandas as pd

from compta_ecom.exporters.excel import (
    export_anomalies_csv_to_bytes,
    export_csv_to_bytes,
    export_to_bytes,
)
from compta_ecom.models import AccountingEntry, Anomaly


def _sample_entries() -> list[AccountingEntry]:
    return [
        AccountingEntry(
            date=datetime.date(2026, 1, 15),
            journal="VE",
            account="411SHOPIFY",
            label="Vente #SH001",
            debit=132.00,
            credit=0.0,
            piece_number="#SH001",
            lettrage="L001",
            channel="shopify",
            entry_type="sale",
        ),
        AccountingEntry(
            date=datetime.date(2026, 1, 15),
            journal="VE",
            account="70701",
            label="CA HT #SH001",
            debit=0.0,
            credit=110.00,
            piece_number="#SH001",
            lettrage="",
            channel="shopify",
            entry_type="sale",
        ),
        AccountingEntry(
            date=datetime.date(2026, 1, 15),
            journal="VE",
            account="44571",
            label="TVA #SH001",
            debit=0.0,
            credit=22.00,
            piece_number="#SH001",
            lettrage="",
            channel="shopify",
            entry_type="sale",
        ),
    ]


def _sample_anomalies() -> list[Anomaly]:
    return [
        Anomaly(
            type="unknown_country",
            severity="warning",
            reference="#SH002",
            channel="shopify",
            detail="Pays inconnu XX",
            expected_value=None,
            actual_value="XX",
        ),
    ]


class TestExportToBytes:
    """Tests pour export_to_bytes()."""

    def test_returns_valid_xlsx(self, sample_config: object) -> None:
        entries = _sample_entries()
        anomalies = _sample_anomalies()
        result = export_to_bytes(entries, anomalies, sample_config)  # type: ignore[arg-type]

        assert isinstance(result, BytesIO)

        wb = openpyxl.load_workbook(result)
        assert "Écritures" in wb.sheetnames
        assert "Anomalies" in wb.sheetnames

        ws_entries = wb["Écritures"]
        # Header + 3 data rows
        assert ws_entries.max_row == 4

        ws_anomalies = wb["Anomalies"]
        # Header + 1 data row
        assert ws_anomalies.max_row == 2

    def test_empty_data(self, sample_config: object) -> None:
        result = export_to_bytes([], [], sample_config)  # type: ignore[arg-type]
        assert isinstance(result, BytesIO)

        wb = openpyxl.load_workbook(result)
        assert "Écritures" in wb.sheetnames
        assert "Anomalies" in wb.sheetnames


class TestExportCsvToBytes:
    """Tests pour export_csv_to_bytes()."""

    def test_returns_valid_csv(self) -> None:
        entries = _sample_entries()
        result = export_csv_to_bytes(entries)

        assert isinstance(result, BytesIO)

        # Check UTF-8 BOM
        raw = result.read()
        assert raw[:3] == b"\xef\xbb\xbf"

        result.seek(0)
        df = pd.read_csv(result, encoding="utf-8-sig", sep=";")
        assert len(df) == 3
        assert list(df.columns) == [
            "date", "journal", "account", "label", "debit", "credit",
            "piece_number", "lettrage", "channel", "entry_type",
        ]

    def test_empty_entries(self) -> None:
        result = export_csv_to_bytes([])
        assert isinstance(result, BytesIO)

        result.seek(0)
        df = pd.read_csv(result, encoding="utf-8-sig", sep=";")
        assert len(df) == 0
        assert "date" in df.columns


class TestExportAnomaliesCsvToBytes:
    """Tests pour export_anomalies_csv_to_bytes()."""

    def test_returns_valid_csv(self) -> None:
        anomalies = _sample_anomalies()
        result = export_anomalies_csv_to_bytes(anomalies)

        assert isinstance(result, BytesIO)

        # Check UTF-8 BOM
        raw = result.read()
        assert raw[:3] == b"\xef\xbb\xbf"

        result.seek(0)
        df = pd.read_csv(result, encoding="utf-8-sig", sep=";")
        assert len(df) == 1
        assert list(df.columns) == [
            "type", "severity", "reference", "channel", "detail",
            "expected_value", "actual_value",
        ]
        assert df.iloc[0]["type"] == "unknown_country"

    def test_empty_anomalies(self) -> None:
        result = export_anomalies_csv_to_bytes([])
        assert isinstance(result, BytesIO)

        result.seek(0)
        df = pd.read_csv(result, encoding="utf-8-sig", sep=";")
        assert len(df) == 0
        assert "type" in df.columns
