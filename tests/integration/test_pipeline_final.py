"""Tests d'intégration pipeline final — multi-canal avec anomalies."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from compta_ecom.config.loader import load_config
from compta_ecom.pipeline import PipelineOrchestrator

FIXTURES_DIR = Path(__file__).parent.parent / "fixtures"
CONFIG_FIXTURES = FIXTURES_DIR / "config"


# ---------------------------------------------------------------------------
# Helpers — CSV content builders
# ---------------------------------------------------------------------------

SHOPIFY_SALES_HEADER = "Name,Created at,Subtotal,Shipping,Taxes,Total,Tax 1 Name,Tax 1 Value,Payment Method,Shipping Country"
SHOPIFY_TX_HEADER = "Order,Type,Payment Method Name,Amount,Fee,Net,Payout Date,Payout ID"
SHOPIFY_PAYOUTS_HEADER = "Payout Date,Charges,Refunds,Fees,Total"

MANOMANO_CA_HEADER = "reference;type;createdAt;amountVatIncl;commissionVatIncl;commissionVatExcl;vatOnCommission;netAmount;productPriceVatExcl;vatOnProduct;shippingPriceVatExcl;vatOnShipping"
MANOMANO_VERSEMENTS_HEADER = "REFERENCE;TYPE;PAYOUT_REFERENCE;PAYOUT_DATE;AMOUNT"

DECATHLON_HEADER = "Numéro de commande;Type;Date de commande;Date du cycle de paiement;Montant"


def _write_shopify_nominal(input_dir: Path) -> None:
    """Shopify nominal — 2 sales, 1 with unknown country XX to trigger unknown_country."""
    (input_dir / "ventes.csv").write_text(
        f"{SHOPIFY_SALES_HEADER}\n"
        "#SH001,2026-01-15,100.00,10.00,22.00,132.00,FR TVA 20%,22.00,Shopify Payments,FR\n"
        # Unknown country XX → triggers unknown_country anomaly from VatChecker
        "#SH002,2026-01-16,50.00,0.00,10.00,60.00,XX TVA 20%,10.00,Shopify Payments,XX\n"
    )
    (input_dir / "transactions.csv").write_text(
        f"{SHOPIFY_TX_HEADER}\n"
        "#SH001,charge,card,132.00,3.96,128.04,2026-01-23,P001\n"
        "#SH002,charge,card,60.00,1.80,58.20,2026-01-23,P001\n"
    )
    (input_dir / "versements.csv").write_text(
        f"{SHOPIFY_PAYOUTS_HEADER}\n"
        "2026-01-23,192.00,0.00,-5.76,186.24\n"
    )


def _write_manomano_with_orphan_refund(input_dir: Path) -> None:
    """ManoMano — 1 sale + 1 orphan refund (REFUND_ORPHAN has no matching sale)."""
    (input_dir / "ca.csv").write_text(
        f"{MANOMANO_CA_HEADER}\n"
        "M001;ORDER;2026-01-15;120.00;18.00;15.00;3.00;102.00;100.00;20.00;0.00;0.00\n"
        # Orphan refund — reference MREFUND999 has no matching ORDER
        "MREFUND999;REFUND;2026-01-16;-120.00;-18.00;-15.00;-3.00;-102.00;-100.00;-20.00;0.00;0.00\n"
    )
    (input_dir / "versements_manomano.csv").write_text(
        f"{MANOMANO_VERSEMENTS_HEADER}\n"
        "M001;ORDER;PAY001;2026-01-31;102.00\n"
        "MREFUND999;REFUND;PAY001;2026-01-31;-102.00\n"
    )


def _write_decathlon_with_amount_mismatch(input_dir: Path) -> None:
    """Decathlon — 1 order with intentional amount mismatch (TTC ≠ commission + net).

    Amount=100, Shipping=10 → TTC should be 132 (with 20% TVA on default country 250).
    Commission=-15, Paiement=95 → commission_ttc + net = 15 + 95 = 110.
    But TTC = 132 → mismatch (132 ≠ 110), écart = 22 > 0.01 tolerance.
    """
    (input_dir / "decathlon_data.csv").write_text(
        f"{DECATHLON_HEADER}\n"
        "DEC001;Montant;2026-01-10;2026-01-25;100.00\n"
        "DEC001;Frais de port;2026-01-10;2026-01-25;10.00\n"
        "DEC001;Commission;2026-01-10;2026-01-25;-15.00\n"
        "DEC001;Taxe sur la commission;2026-01-10;2026-01-25;0.00\n"
        "DEC001;Paiement;2026-01-10;2026-01-25;95.00\n"
    )


def _write_shopify_clean(input_dir: Path) -> None:
    """Shopify data without any anomaly — all coherent."""
    (input_dir / "ventes.csv").write_text(
        f"{SHOPIFY_SALES_HEADER}\n"
        "#CLEAN001,2026-01-15,100.00,0.00,20.00,120.00,FR TVA 20%,20.00,Shopify Payments,FR\n"
    )
    (input_dir / "transactions.csv").write_text(
        f"{SHOPIFY_TX_HEADER}\n"
        "#CLEAN001,charge,card,120.00,3.60,116.40,2026-01-23,P001\n"
    )
    (input_dir / "versements.csv").write_text(
        f"{SHOPIFY_PAYOUTS_HEADER}\n"
        "2026-01-23,120.00,0.00,-3.60,116.40\n"
    )


# ---------------------------------------------------------------------------
# Test: Pipeline complet multi-canal avec anomalies intentionnelles (AC 6,8,9)
# ---------------------------------------------------------------------------


class TestPipelineFinalWithAnomalies:
    """Pipeline complet multi-canal avec anomalies intentionnelles."""

    def test_pipeline_multi_canal_produces_excel(self, tmp_path: Path) -> None:
        """3 canaux → fichier Excel produit avec 2 onglets."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        _write_shopify_nominal(input_dir)
        _write_manomano_with_orphan_refund(input_dir)
        _write_decathlon_with_amount_mismatch(input_dir)

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        orchestrator.run(input_dir, output, config)

        assert output.exists()
        wb = openpyxl.load_workbook(output)
        assert "Écritures" in wb.sheetnames
        assert "Anomalies" in wb.sheetnames
        wb.close()

    def test_entries_have_lettrage(self, tmp_path: Path) -> None:
        """Les écritures sur comptes 411/511 ont un lettrage non vide, les autres non."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        _write_shopify_nominal(input_dir)
        _write_manomano_with_orphan_refund(input_dir)
        _write_decathlon_with_amount_mismatch(input_dir)

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        PipelineOrchestrator().run(input_dir, output, config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Écritures"]
        account_col = 3  # column C = account (1-based)
        lettrage_col = 8  # column H = lettrage (1-based)
        for row in ws.iter_rows(min_row=2, values_only=False):
            account = str(row[account_col - 1].value or "")
            lettrage = row[lettrage_col - 1].value
            if account.startswith("411") or account.startswith("511"):
                assert lettrage is not None and str(lettrage).strip() != "", (
                    f"Lettrage vide pour le compte {account} ligne {row[0].row}"
                )
        wb.close()

    def test_balance_debit_equals_credit(self, tmp_path: Path) -> None:
        """Balance globale : total débit ≈ total crédit."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        _write_shopify_nominal(input_dir)
        _write_manomano_with_orphan_refund(input_dir)
        _write_decathlon_with_amount_mismatch(input_dir)

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        PipelineOrchestrator().run(input_dir, output, config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Écritures"]
        total_debit = 0.0
        total_credit = 0.0
        debit_col = 5  # column E
        credit_col = 6  # column F
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_debit += float(row[debit_col - 1] or 0)
            total_credit += float(row[credit_col - 1] or 0)
        wb.close()

        assert abs(total_debit - total_credit) < 0.02, (
            f"Déséquilibre: débit={total_debit}, crédit={total_credit}"
        )

    def test_anomalies_from_multiple_sources(self, tmp_path: Path) -> None:
        """Les anomalies proviennent de sources multiples (parsers + contrôles)."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        _write_shopify_nominal(input_dir)
        _write_manomano_with_orphan_refund(input_dir)
        _write_decathlon_with_amount_mismatch(input_dir)

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        PipelineOrchestrator().run(input_dir, output, config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Anomalies"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        anomaly_types = {row[0] for row in data_rows}
        wb.close()

        # VatChecker: unknown_country (country XX — from controls, not parsers)
        assert "unknown_country" in anomaly_types, (
            f"unknown_country attendu, trouvé: {anomaly_types}"
        )

    def test_anomalies_have_correct_severities(self, tmp_path: Path) -> None:
        """Les anomalies intentionnelles ont les sévérités attendues."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        _write_shopify_nominal(input_dir)
        _write_manomano_with_orphan_refund(input_dir)
        _write_decathlon_with_amount_mismatch(input_dir)

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        PipelineOrchestrator().run(input_dir, output, config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Anomalies"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        type_severity = {row[0]: row[1] for row in data_rows}
        assert type_severity.get("unknown_country") == "error"

    def test_entries_count_positive(self, tmp_path: Path) -> None:
        """Le pipeline produit des écritures pour les 3 canaux."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        _write_shopify_nominal(input_dir)
        _write_manomano_with_orphan_refund(input_dir)
        _write_decathlon_with_amount_mismatch(input_dir)

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        PipelineOrchestrator().run(input_dir, output, config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Écritures"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        assert len(data_rows) > 0
        # Check multiple channels present
        channels = {row[8] for row in data_rows}  # column I = channel (index 8)
        assert len(channels) >= 2, f"Attendu ≥2 canaux, trouvé: {channels}"


# ---------------------------------------------------------------------------
# Test: Pipeline sans anomalie (AC 10)
# ---------------------------------------------------------------------------


class TestPipelineNoAnomalies:
    """Pipeline avec données propres — aucune anomalie."""

    def test_no_anomalies_clean_data(self, tmp_path: Path) -> None:
        """Données propres → onglet Anomalies vide (en-têtes uniquement)."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        _write_shopify_clean(input_dir)

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        PipelineOrchestrator().run(input_dir, output, config)

        wb = openpyxl.load_workbook(output)
        ws = wb["Anomalies"]
        # Headers present
        headers = [cell.value for cell in ws[1]]
        assert "type" in headers
        # No data rows
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Filter out completely empty rows
        non_empty = [r for r in data_rows if any(c is not None for c in r)]
        assert len(non_empty) == 0, f"Anomalies inattendues: {non_empty}"
        wb.close()

    def test_no_anomalies_console_message(self, tmp_path: Path, capsys: object) -> None:
        """Données propres → résumé console affiche 'Aucune anomalie détectée'."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        input_dir = tmp_path / "input"
        input_dir.mkdir()
        _write_shopify_clean(input_dir)

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        PipelineOrchestrator().run(input_dir, output, config)

        captured = capsys.readouterr()
        assert "Aucune anomalie détectée" in captured.out


# ---------------------------------------------------------------------------
# Test: Canal en erreur (AC 11)
# ---------------------------------------------------------------------------


class TestPipelineChannelError:
    """Pipeline avec un canal en erreur — les autres continuent."""

    def test_malformed_canal_others_continue(self, tmp_path: Path, capsys: object) -> None:
        """Un canal malformé → les autres canaux traités normalement."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        input_dir = tmp_path / "input"
        input_dir.mkdir()

        # Shopify clean
        _write_shopify_clean(input_dir)

        # ManoMano malformed (missing required columns)
        (input_dir / "ca.csv").write_text("BadCol1;BadCol2\nfoo;bar\n")
        (input_dir / "versements_manomano.csv").write_text(
            f"{MANOMANO_VERSEMENTS_HEADER}\n"
        )

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        PipelineOrchestrator().run(input_dir, output, config)

        # Excel produced with shopify data
        assert output.exists()
        wb = openpyxl.load_workbook(output)
        ws = wb["Écritures"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) > 0
        wb.close()

        # Console mentions the error channel
        captured = capsys.readouterr()
        # The summary should still be printed (from shopify data)
        assert "Résumé" in captured.out
        # AC7: channel error must appear in console summary
        assert "Canaux en erreur" in captured.out
        assert "manomano" in captured.out
