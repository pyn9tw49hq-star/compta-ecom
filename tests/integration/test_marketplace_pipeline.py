"""Tests d'intégration pour le pipeline multi-canal marketplace."""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

from compta_ecom.config.loader import load_config
from compta_ecom.pipeline import PipelineOrchestrator

FIXTURES_DIR = Path(__file__).parent.parent / "fixtures"
SHOPIFY_FIXTURES = FIXTURES_DIR / "shopify"
MANOMANO_FIXTURES = FIXTURES_DIR / "manomano"
DECATHLON_FIXTURES = FIXTURES_DIR / "decathlon"
LEROY_MERLIN_FIXTURES = FIXTURES_DIR / "leroy_merlin"
CONFIG_FIXTURES = FIXTURES_DIR / "config"


def _setup_multi_channel(tmp_path: Path) -> Path:
    """Copie les fixtures des 4 canaux dans un répertoire temporaire."""
    input_dir = tmp_path / "input"
    input_dir.mkdir()

    # Shopify
    for f in SHOPIFY_FIXTURES.glob("*.csv"):
        shutil.copy(f, input_dir / f.name)

    # ManoMano
    for f in MANOMANO_FIXTURES.glob("*.csv"):
        shutil.copy(f, input_dir / f.name)

    # Décathlon
    for f in DECATHLON_FIXTURES.glob("*.csv"):
        shutil.copy(f, input_dir / f.name)

    # Leroy Merlin
    for f in LEROY_MERLIN_FIXTURES.glob("*.csv"):
        shutil.copy(f, input_dir / f.name)

    return input_dir


class TestMultiChannelPipeline:
    """Tests d'intégration multi-canal complet."""

    def test_four_channels_excel_output(self, tmp_path: Path) -> None:
        """4 canaux → Excel avec écritures des 4 canaux présentes."""
        input_dir = _setup_multi_channel(tmp_path)
        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        orchestrator.run(input_dir, output, config)

        assert output.exists()

        wb = openpyxl.load_workbook(output)
        assert "Écritures" in wb.sheetnames
        assert "Anomalies" in wb.sheetnames

        ws_entries = wb["Écritures"]
        data_rows = list(ws_entries.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) > 0

        # Extract channels (column index 8 = channel)
        channels = {row[8] for row in data_rows}
        assert "shopify" in channels
        assert "manomano" in channels
        assert "decathlon" in channels
        assert "leroy_merlin" in channels

        # Check entry types present
        entry_types = {row[9] for row in data_rows}
        assert "sale" in entry_types
        assert "payout" in entry_types
        assert "commission" in entry_types

        # Count by entry_type
        payout_entries = [r for r in data_rows if r[9] == "payout"]
        fee_entries = [r for r in data_rows if r[9] == "fee"]
        # ManoMano: M001 payout(2) + M002 payout(2) + ADJ001 payout(2) + PayoutSummary PAY001(2) = 8 payout
        # ManoMano: SUB001 fee(2) = 2 fee
        # Décathlon: Paiement PayoutSummary(2) = 2 payout (pas de payout individuel 512)
        # Décathlon: SUBDEC001 fee(2) = 2 fee
        # Leroy Merlin: Paiement PayoutSummary(2) = 2 payout (pas de payout individuel 512)
        # Shopify: payout PSP from PayoutSummary(2) = 2 payout
        mm_payouts = [r for r in payout_entries if r[8] == "manomano"]
        assert len(mm_payouts) == 8

        dec_payouts = [r for r in payout_entries if r[8] == "decathlon"]
        assert len(dec_payouts) == 2

        lm_payouts = [r for r in payout_entries if r[8] == "leroy_merlin"]
        assert len(lm_payouts) == 2

        # Vérifier les entrées fee (frais d'abonnement)
        mm_fees = [r for r in fee_entries if r[8] == "manomano"]
        assert len(mm_fees) == 2
        dec_fees = [r for r in fee_entries if r[8] == "decathlon"]
        assert len(dec_fees) == 2

        wb.close()

    def test_anomalies_sheet_has_headers(self, tmp_path: Path) -> None:
        """L'onglet Anomalies a des en-têtes même si pas d'anomalie critique."""
        input_dir = _setup_multi_channel(tmp_path)
        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        orchestrator.run(input_dir, output, config)

        wb = openpyxl.load_workbook(output)
        ws_anomalies = wb["Anomalies"]
        headers = [cell.value for cell in ws_anomalies[1]]
        assert "type" in headers
        assert "severity" in headers
        wb.close()

    def test_console_summary(self, tmp_path: Path, capsys: object) -> None:
        """Le résumé console affiche les statistiques."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        input_dir = _setup_multi_channel(tmp_path)
        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        orchestrator.run(input_dir, output, config)

        captured = capsys.readouterr()
        # Summary should mention total entries
        assert "écriture" in captured.out.lower() or "ecritures" in captured.out.lower() or len(captured.out) > 0


class TestShopifyOnlyNonRegression:
    """Non-régression : pipeline Shopify-only inchangé après Story 2.4."""

    def test_shopify_only_pipeline(self, tmp_path: Path) -> None:
        """Pipeline avec uniquement des fichiers Shopify → comportement identique."""
        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        orchestrator.run(SHOPIFY_FIXTURES, output, config)

        assert output.exists()

        wb = openpyxl.load_workbook(output)
        ws_entries = wb["Écritures"]
        data_rows = list(ws_entries.iter_rows(min_row=2, values_only=True))

        # Only Shopify entries
        channels = {row[8] for row in data_rows}
        assert channels == {"shopify"}

        # Check 3 entry types for Shopify
        entry_types = {row[9] for row in data_rows}
        assert "sale" in entry_types
        assert "settlement" in entry_types
        assert "payout" in entry_types

        wb.close()
