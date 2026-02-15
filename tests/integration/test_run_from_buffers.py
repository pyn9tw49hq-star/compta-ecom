"""Tests d'intégration — run_from_buffers() vs run()."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from compta_ecom.config.loader import load_config
from compta_ecom.pipeline import PipelineOrchestrator

FIXTURES_DIR = Path(__file__).parent.parent / "fixtures"
CONFIG_FIXTURES = FIXTURES_DIR / "config"

# ---------------------------------------------------------------------------
# CSV content matching test fixture channels.yaml patterns
# ---------------------------------------------------------------------------

SHOPIFY_SALES_HEADER = "Name,Created at,Subtotal,Shipping,Taxes,Total,Tax 1 Name,Tax 1 Value,Payment Method,Shipping Country"
SHOPIFY_TX_HEADER = "Order,Type,Payment Method Name,Amount,Fee,Net,Payout Date,Payout ID"
SHOPIFY_PAYOUTS_HEADER = "Payout Date,Charges,Refunds,Fees,Total"

MANOMANO_CA_HEADER = "reference;type;createdAt;amountVatIncl;commissionVatIncl;commissionVatExcl;vatOnCommission;netAmount;productPriceVatExcl;vatOnProduct;shippingPriceVatExcl;vatOnShipping"
MANOMANO_VERSEMENTS_HEADER = "REFERENCE;TYPE;PAYOUT_REFERENCE;PAYOUT_DATE;AMOUNT"
MANOMANO_OD_HEADER = "Order Reference;Billing Country ISO"

DECATHLON_HEADER = "Numéro de commande;Type;Date de commande;Date du cycle de paiement;Montant"


def _build_test_data() -> dict[str, str]:
    """Returns {filename: csv_content} for a multi-channel test scenario."""
    return {
        "ventes.csv": (
            f"{SHOPIFY_SALES_HEADER}\n"
            "#SH001,2026-01-15,100.00,10.00,22.00,132.00,FR TVA 20%,22.00,Shopify Payments,FR\n"
            "#SH002,2026-01-16,50.00,0.00,10.00,60.00,FR TVA 20%,10.00,Shopify Payments,FR\n"
        ),
        "transactions.csv": (
            f"{SHOPIFY_TX_HEADER}\n"
            "#SH001,charge,card,132.00,3.96,128.04,2026-01-23,P001\n"
            "#SH002,charge,card,60.00,1.80,58.20,2026-01-23,P001\n"
        ),
        "versements.csv": (
            f"{SHOPIFY_PAYOUTS_HEADER}\n"
            "2026-01-23,192.00,0.00,-5.76,186.24\n"
        ),
        "ca.csv": (
            f"{MANOMANO_CA_HEADER}\n"
            "M001;ORDER;2026-01-15;120.00;-18.00;-15.00;-3.00;102.00;100.00;20.00;0.00;0.00\n"
        ),
        "versements_manomano.csv": (
            f"{MANOMANO_VERSEMENTS_HEADER}\n"
            "M001;ORDER;PAY001;2026-01-31;102.00\n"
        ),
        "commandes_manomano.csv": (
            f"{MANOMANO_OD_HEADER}\n"
            "M001;FR\n"
        ),
        "decathlon_data.csv": (
            f"{DECATHLON_HEADER}\n"
            "DEC001;Montant;2026-01-10;2026-01-25;100.00\n"
            "DEC001;Frais de port;2026-01-10;2026-01-25;10.00\n"
            "DEC001;Commission;2026-01-10;2026-01-25;-15.00\n"
            "DEC001;Taxe sur la commission;2026-01-10;2026-01-25;0.00\n"
            "DEC001;Paiement;2026-01-10;2026-01-25;95.00\n"
        ),
    }


class TestRunFromBuffersVsRun:
    """Compare run_from_buffers() with run() on identical data."""

    def _run_both(self, tmp_path: Path) -> tuple[
        list[object], list[object], dict[str, object],
        Path,
    ]:
        """Helper: run both modes, return buffers results + disk output path."""
        test_data = _build_test_data()
        config = load_config(CONFIG_FIXTURES)
        orchestrator = PipelineOrchestrator()

        # --- Disk mode ---
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        for filename, content in test_data.items():
            (input_dir / filename).write_text(content, encoding="utf-8")
        output_path = tmp_path / "output.xlsx"
        orchestrator.run(input_dir, output_path, config)

        # --- Buffer mode ---
        files_bytes: dict[str, bytes] = {
            name: content.encode("utf-8") for name, content in test_data.items()
        }
        entries_buf, anomalies_buf, summary, _txs = orchestrator.run_from_buffers(files_bytes, config)

        return entries_buf, anomalies_buf, summary, output_path

    def test_same_entry_count(self, tmp_path: Path) -> None:
        """run_from_buffers() produces the same number of entries as run()."""
        entries_buf, _, _, output_path = self._run_both(tmp_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Écritures"]
        disk_count = ws.max_row - 1  # minus header
        wb.close()

        assert len(entries_buf) == disk_count, (
            f"Entries: buffer={len(entries_buf)} vs disk={disk_count}"
        )

    def test_same_anomaly_count(self, tmp_path: Path) -> None:
        """run_from_buffers() produces the same number of anomalies as run()."""
        _, anomalies_buf, _, output_path = self._run_both(tmp_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Anomalies"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        disk_count = len([r for r in rows if any(c is not None for c in r)])
        wb.close()

        assert len(anomalies_buf) == disk_count, (
            f"Anomalies: buffer={len(anomalies_buf)} vs disk={disk_count}"
        )

    def test_debit_credit_balance(self, tmp_path: Path) -> None:
        """Total debit == total credit in buffer mode."""
        entries_buf, _, _, _ = self._run_both(tmp_path)

        total_debit = round(sum(e.debit for e in entries_buf), 2)
        total_credit = round(sum(e.credit for e in entries_buf), 2)
        assert abs(total_debit - total_credit) < 0.02, (
            f"Déséquilibre buffer: débit={total_debit}, crédit={total_credit}"
        )

    def test_same_totals_as_disk(self, tmp_path: Path) -> None:
        """Buffer mode totals match disk mode totals."""
        entries_buf, _, _, output_path = self._run_both(tmp_path)

        # Buffer totals
        buf_debit = round(sum(e.debit for e in entries_buf), 2)
        buf_credit = round(sum(e.credit for e in entries_buf), 2)

        # Disk totals
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Écritures"]
        disk_debit = 0.0
        disk_credit = 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            disk_debit += float(row[4] or 0)  # debit col (index 4)
            disk_credit += float(row[5] or 0)  # credit col (index 5)
        wb.close()
        disk_debit = round(disk_debit, 2)
        disk_credit = round(disk_credit, 2)

        assert buf_debit == disk_debit, f"Debit: buf={buf_debit} vs disk={disk_debit}"
        assert buf_credit == disk_credit, f"Credit: buf={buf_credit} vs disk={disk_credit}"

    def test_summary_keys_present(self, tmp_path: Path) -> None:
        """Summary dict has all required keys."""
        _, _, summary, _ = self._run_both(tmp_path)

        assert "transactions_par_canal" in summary
        assert "ecritures_par_type" in summary
        assert "totaux" in summary

        totaux = summary["totaux"]
        assert "debit" in totaux
        assert "credit" in totaux

    def test_summary_values_coherent(self, tmp_path: Path) -> None:
        """Summary values match the actual entries."""
        entries_buf, _, summary, _ = self._run_both(tmp_path)

        # Totals match
        expected_debit = round(sum(e.debit for e in entries_buf), 2)
        expected_credit = round(sum(e.credit for e in entries_buf), 2)
        assert summary["totaux"]["debit"] == expected_debit
        assert summary["totaux"]["credit"] == expected_credit

        # Entry type counts match
        from collections import Counter
        expected_types = dict(Counter(e.entry_type for e in entries_buf))
        assert summary["ecritures_par_type"] == expected_types

    def test_multiple_channels_in_summary(self, tmp_path: Path) -> None:
        """Summary includes transactions from multiple channels."""
        _, _, summary, _ = self._run_both(tmp_path)

        tx_par_canal = summary["transactions_par_canal"]
        assert len(tx_par_canal) >= 2, f"Expected >=2 channels, got: {tx_par_canal}"
