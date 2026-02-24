"""Tests d'intégration pour le pipeline Shopify complet."""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pytest

from compta_ecom.config.loader import load_config
from compta_ecom.main import main
from compta_ecom.models import NoResultError
from compta_ecom.pipeline import PipelineOrchestrator

FIXTURES_DIR = Path(__file__).parent.parent / "fixtures"
SHOPIFY_FIXTURES = FIXTURES_DIR / "shopify"
CONFIG_FIXTURES = FIXTURES_DIR / "config"


class TestPipelineNominal:
    """Pipeline nominal : 3 CSV → Excel avec 3 types d'écritures."""

    def test_nominal_excel_output(self, tmp_path: Path) -> None:
        """3 fichiers CSV Shopify → Excel avec 2 onglets, écritures des 3 types."""
        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        orchestrator.run(SHOPIFY_FIXTURES, output, config)

        assert output.exists()

        wb = openpyxl.load_workbook(output)
        assert "Écritures" in wb.sheetnames
        assert "Anomalies" in wb.sheetnames

        ws_entries = wb["Écritures"]
        data_rows = list(ws_entries.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) > 0

        # Check 3 entry types present
        entry_types = {row[9] for row in data_rows}  # entry_type is column 10 (index 9)
        assert "sale" in entry_types
        assert "settlement" in entry_types
        assert "payout" in entry_types

        wb.close()

    def test_anomalies_sheet_has_headers(self, tmp_path: Path) -> None:
        """L'onglet Anomalies a des en-têtes même si aucune anomalie critique."""
        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        orchestrator.run(SHOPIFY_FIXTURES, output, config)

        wb = openpyxl.load_workbook(output)
        ws_anomalies = wb["Anomalies"]
        headers = [cell.value for cell in ws_anomalies[1]]
        assert "type" in headers
        assert "severity" in headers
        assert "detail" in headers
        wb.close()


class TestPipelineWithAnomalies:
    """Pipeline avec anomalies : commande orpheline."""

    def test_orphan_sale_anomaly(self, tmp_path: Path) -> None:
        """CSV ventes avec une commande orpheline → onglet Anomalies contient orphan_sale_summary."""
        # Create modified sales CSV with an extra order not in transactions
        input_dir = tmp_path / "input"
        input_dir.mkdir()

        # Copy transactions and versements as-is
        shutil.copy(SHOPIFY_FIXTURES / "transactions.csv", input_dir / "transactions.csv")
        shutil.copy(SHOPIFY_FIXTURES / "versements.csv", input_dir / "versements.csv")

        # Create modified sales with an orphan order
        sales_content = (SHOPIFY_FIXTURES / "ventes.csv").read_text()
        sales_content += '#ORPHAN99,2026-01-18,30.00,0.00,6.00,36.00,FR TVA 20%,6.00,Shopify Payments,FR\n'
        (input_dir / "ventes.csv").write_text(sales_content)

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        orchestrator.run(input_dir, output, config)

        wb = openpyxl.load_workbook(output)
        ws_anomalies = wb["Anomalies"]
        data_rows = list(ws_anomalies.iter_rows(min_row=2, values_only=True))
        anomaly_types = [row[0] for row in data_rows]
        assert "orphan_sale_summary" in anomaly_types
        wb.close()


class TestPipelineNoFiles:
    """Pipeline sans fichiers : répertoire vide → aucun résultat."""

    def test_empty_directory_raises(self, tmp_path: Path) -> None:
        """Répertoire vide → NoResultError."""
        input_dir = tmp_path / "empty_input"
        input_dir.mkdir()
        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        with pytest.raises(NoResultError):
            orchestrator.run(input_dir, output, config)

        assert not output.exists()


class TestPipelineChannelError:
    """Pipeline canal en erreur : CSV malformé."""

    def test_malformed_csv_parse_error(self, tmp_path: Path, capsys: object) -> None:
        """CSV malformé (colonne manquante) → ParseError catchée, résumé mentionne l'erreur."""
        import _pytest.capture

        assert isinstance(capsys, _pytest.capture.CaptureFixture)

        input_dir = tmp_path / "input"
        input_dir.mkdir()

        # Create a malformed sales CSV (missing required columns)
        (input_dir / "ventes.csv").write_text("BadCol1,BadCol2\nfoo,bar\n")
        # Provide valid transactions + versements so that parse error is on sales only
        shutil.copy(SHOPIFY_FIXTURES / "transactions.csv", input_dir / "transactions.csv")
        shutil.copy(SHOPIFY_FIXTURES / "versements.csv", input_dir / "versements.csv")

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        # sales file is malformed → ParseError → caught → NoResultError (only 1 channel)
        with pytest.raises(NoResultError):
            orchestrator.run(input_dir, output, config)


class TestCLIExitCodes:
    """Tests des exit codes CLI."""

    def test_exit_code_0_nominal(self, tmp_path: Path) -> None:
        """Pipeline nominal via main() → exit code 0 (pas de SystemExit)."""
        output = tmp_path / "output.xlsx"
        main([
            str(SHOPIFY_FIXTURES),
            str(output),
            "--config-dir",
            str(CONFIG_FIXTURES),
            "--log-level",
            "WARNING",
        ])
        assert output.exists()

    def test_exit_code_3_no_results(self, tmp_path: Path) -> None:
        """Répertoire vide via main() → exit code 3."""
        input_dir = tmp_path / "empty"
        input_dir.mkdir()
        output = tmp_path / "output.xlsx"

        with pytest.raises(SystemExit) as exc_info:
            main([
                str(input_dir),
                str(output),
                "--config-dir",
                str(CONFIG_FIXTURES),
                "--log-level",
                "WARNING",
            ])
        assert exc_info.value.code == 3


class TestPipelineDirectPayment:
    """Pipeline avec paiement direct Klarna → écritures VE + RG, anomalie direct_payment."""

    def test_klarna_direct_payment_end_to_end(self, tmp_path: Path) -> None:
        """Vente Klarna orpheline → VE (3 lignes) + RG (2 lignes), lettrage 411 cohérent."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()

        # Copy transactions and versements as-is
        shutil.copy(SHOPIFY_FIXTURES / "transactions.csv", input_dir / "transactions.csv")
        shutil.copy(SHOPIFY_FIXTURES / "versements.csv", input_dir / "versements.csv")

        # Create modified sales with a Klarna orphan order
        sales_content = (SHOPIFY_FIXTURES / "ventes.csv").read_text()
        sales_content += '#KLARNA01,2026-01-20,80.00,0.00,16.00,96.00,FR TVA 20%,16.00,Klarna,FR\n'
        (input_dir / "ventes.csv").write_text(sales_content)

        config = load_config(CONFIG_FIXTURES)
        output = tmp_path / "output.xlsx"

        orchestrator = PipelineOrchestrator()
        orchestrator.run(input_dir, output, config)

        wb = openpyxl.load_workbook(output)
        ws_entries = wb["Écritures"]
        headers = [cell.value for cell in ws_entries[1]]
        col = {name: idx for idx, name in enumerate(headers)}
        data_rows = list(ws_entries.iter_rows(min_row=2, values_only=True))

        # Filter entries for the Klarna order
        klarna_entries = [r for r in data_rows if r[col["piece_number"]] == "#KLARNA01"]

        # VE entries: 411 D + 707 C + 4457 C = 3 lines
        ve_entries = [r for r in klarna_entries if r[col["journal"]] == "VE"]
        assert len(ve_entries) == 3, f"Attendu 3 écritures VE, trouvé {len(ve_entries)}"

        # RG entries: 46740000 D + 411 C = 2 lines
        rg_entries = [r for r in klarna_entries if r[col["journal"]] == "RG"]
        assert len(rg_entries) == 2, f"Attendu 2 écritures RG, trouvé {len(rg_entries)}"

        # RG debit on 46740000
        rg_debit = [r for r in rg_entries if (r[col["debit"]] or 0) > 0]
        assert len(rg_debit) == 1
        assert rg_debit[0][col["account"]] == "46740000"
        assert rg_debit[0][col["debit"]] == 96.0

        # RG credit on 411SHOPIFY
        rg_credit = [r for r in rg_entries if (r[col["credit"]] or 0) > 0]
        assert len(rg_credit) == 1
        assert rg_credit[0][col["account"]] == "411SHOPIFY"
        assert rg_credit[0][col["credit"]] == 96.0

        # Lettrage 411 cohérent entre VE et RG (même code alphabétique)
        ve_411 = [r for r in ve_entries if r[col["account"]] == "411SHOPIFY"]
        rg_411 = [r for r in rg_entries if r[col["account"]] == "411SHOPIFY"]
        assert len(ve_411) == 1
        assert len(rg_411) == 1
        assert ve_411[0][col["lettrage"]] == rg_411[0][col["lettrage"]]
        assert ve_411[0][col["lettrage"]] not in (None, "")

        # Anomalie direct_payment présente, pas orphan_sale_summary mentionnant #KLARNA01
        ws_anomalies = wb["Anomalies"]
        anomaly_rows = list(ws_anomalies.iter_rows(min_row=2, values_only=True))
        anomaly_types_for_klarna = [
            r[0] for r in anomaly_rows if r[2] == "#KLARNA01"  # type, severity, reference
        ]
        assert "direct_payment" in anomaly_types_for_klarna
        # orphan_sale_summary should not mention #KLARNA01 since it's a direct_payment
        orphan_summaries = [r for r in anomaly_rows if r[0] == "orphan_sale_summary"]
        for summary_row in orphan_summaries:
            assert "#KLARNA01" not in (summary_row[3] or "")

        wb.close()


DETAIL_FIXTURES = SHOPIFY_FIXTURES / "detail_versements"


class TestPipelineDetailedLettrage:
    """AC16 : Pipeline avec detail_versements → 1 seule paire payout globale par versement,
    lettrage = payout_reference partagé entre settlement et payout sur 511."""

    def _run_pipeline_with_details(self, tmp_path: Path) -> list[tuple]:
        """Helper : exécute le pipeline avec detail_versements et retourne (headers, data_rows)."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        shutil.copy(SHOPIFY_FIXTURES / "ventes.csv", input_dir / "ventes.csv")
        shutil.copy(SHOPIFY_FIXTURES / "transactions.csv", input_dir / "transactions.csv")
        shutil.copy(SHOPIFY_FIXTURES / "versements.csv", input_dir / "versements.csv")

        # Copy detail_versements files
        for detail_file in DETAIL_FIXTURES.glob("*.csv"):
            shutil.copy(detail_file, input_dir / detail_file.name)

        # Load config and add payout_details to shopify channel
        config = load_config(CONFIG_FIXTURES)
        config.channels["shopify"].files["payout_details"] = "detail_*.csv"
        config.channels["shopify"].multi_files = ["payout_details"]

        output = tmp_path / "output.xlsx"
        orchestrator = PipelineOrchestrator()
        orchestrator.run(input_dir, output, config)

        assert output.exists()

        wb = openpyxl.load_workbook(output)
        ws = wb["Écritures"]
        headers = [cell.value for cell in ws[1]]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        return headers, data_rows

    def test_one_payout_pair_per_versement(self, tmp_path: Path) -> None:
        """Avec detail files, 1 seule paire payout (2 lignes : 580 D / 46710001 C) par versement."""
        headers, data_rows = self._run_pipeline_with_details(tmp_path)
        col = {name: idx for idx, name in enumerate(headers)}

        payout_entries = [r for r in data_rows if r[col["entry_type"]] == "payout"]

        # 2 versements (P001, P002) → 2 paires = 4 lignes payout
        assert len(payout_entries) == 4, (
            f"Attendu 4 lignes payout (2 paires), trouvé {len(payout_entries)}"
        )

        # Each versement has exactly 1 transit (580) + 1 intermédiaire (46710001) line
        payout_580 = [r for r in payout_entries if str(r[col["account"]]).startswith("580")]
        payout_4671 = [r for r in payout_entries if str(r[col["account"]]) == "46710001"]
        assert len(payout_580) == 2, "Attendu 2 lignes 580 (1 par versement)"
        assert len(payout_4671) == 2, "Attendu 2 lignes 46710001 (1 par versement)"

    def test_lettrage_shared_between_settlement_and_payout(self, tmp_path: Path) -> None:
        """Le lettrage sur 46710001 est partagé entre settlement et payout (même code par versement)."""
        headers, data_rows = self._run_pipeline_with_details(tmp_path)
        col = {name: idx for idx, name in enumerate(headers)}

        # Collect payout entries on 46710001
        payout_4671 = [
            r for r in data_rows
            if r[col["entry_type"]] == "payout"
            and str(r[col["account"]]) == "46710001"
        ]
        payout_lettrage = {r[col["lettrage"]] for r in payout_4671}

        # Collect settlement entries on 46710001
        settlement_4671 = [
            r for r in data_rows
            if r[col["entry_type"]] in ("settlement", "commission")
            and str(r[col["account"]]) == "46710001"
        ]
        settlement_lettrage = {r[col["lettrage"]] for r in settlement_4671 if r[col["lettrage"]]}

        # Payout lettrage codes must be a subset of settlement lettrage codes
        # (each payout shares its lettrage with at least one settlement)
        assert len(payout_lettrage) == 2, f"Attendu 2 codes lettrage payout, trouvé {payout_lettrage}"
        assert payout_lettrage <= settlement_lettrage, (
            f"Lettrage payout {payout_lettrage} non partagé avec settlement {settlement_lettrage}"
        )

    def test_lettrage_balance_on_intermed(self, tmp_path: Path) -> None:
        """Pour chaque payout_reference, somme D 46710001 = somme C 46710001 (settlements + payout)."""
        headers, data_rows = self._run_pipeline_with_details(tmp_path)
        col = {name: idx for idx, name in enumerate(headers)}

        # Collect all 46710001 entries by lettrage
        entries_4671_by_lettrage: dict[str, list[tuple]] = {}
        for row in data_rows:
            account = row[col["account"]]
            lettrage = row[col["lettrage"]]
            if account is None or str(account) != "46710001":
                continue
            if not lettrage:
                continue
            entries_4671_by_lettrage.setdefault(lettrage, []).append(row)

        assert len(entries_4671_by_lettrage) > 0, "No 46710001 entries with lettrage found"

        for lettrage, rows in entries_4671_by_lettrage.items():
            entry_types = {r[col["entry_type"]] for r in rows}
            # settlement or commission entries must exist alongside payout entries
            assert entry_types & {"settlement", "commission"}, f"Lettrage {lettrage}: missing settlement/commission entries"
            assert "payout" in entry_types, f"Lettrage {lettrage}: missing payout entries"

            # Each group must be balanced (sum debits == sum credits)
            total_debit = sum(r[col["debit"]] or 0 for r in rows)
            total_credit = sum(r[col["credit"]] or 0 for r in rows)
            assert round(total_debit, 2) == round(total_credit, 2), (
                f"Lettrage {lettrage}: déséquilibre D={total_debit} C={total_credit}"
            )
